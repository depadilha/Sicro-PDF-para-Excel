import pdfplumber
import os
from openpyxl import Workbook

# Função recebe uma lista, correspondente aos elementos de uma linha do PDF, e um código. Em seguida gera um dicionário
# com as informações buscadas. Dependendo do código fornecido, encaixa cada valor em seu respectivo lugar baseando-se
# na posição (índice) que o valor ocupa na linha.


def lines_to_dic(linha, tipo):
    if tipo == "eq":
        return {"Código": linha[0], "Descrição": " ".join(linha[1:-9]), "Valor de Aquisição": linha[-9], "Depreciação":
                linha[-8], "Oportunidade Capital": linha[-7], "Seguros e Impostos": linha[-6], "Manutenção": linha[-5],
                "Operação": linha[-4], "Mão de Obra de Operação": linha[-3], "Custo Produtivo": linha[-2],
                "Custo Improdutivo": linha[-1]}
    elif tipo == "mo":
        return {"Código": linha[0], "Descrição": " ".join(linha[1:-5]), "Unidade": linha[-5], "Salário": linha[-4],
                "Encargos Totais": linha[-3], "Custo": linha[-2], "Periculosidade": linha[-1]}
    elif tipo == "ma":
        return {"Código": linha[0], "Descrição": " ".join(linha[1:-2]), "Unidade": linha[-2],
                "Preço Unitário": linha[-1]}
    elif tipo == "ccu":
        return {"Código": linha[0], "Descrição": " ".join(linha[1:-2]), "Unidade": linha[-2],
                "Custo Unitário": linha[-1]}

# Função recebe um diretório, que se refere ao nome do arquivo de PDF à ser lido, e um código. A função então abre o
# PDF, cujo o diretório foi fornecido, separa suas páginas em linhas e suas linhas em elementos. Em sequência filtra,
# dependendo do código recebido, as linhas que serão armazenadas. Com a função "lines_to_dic" armazena cada elemento da
# linha em um dicionário diferente, esses dicionários por sua vez são armazenados em quatro listas diferentes,
# dependendo do código fornecido.


def pdf_to_lines(dir_arq, tipo):
    with pdfplumber.open(os.path.join(Dir_PDFs, dir_arq)) as pdf:
        for i, pages in enumerate(pdf.pages):
            page = pdf.pages[i]  # Page
            line = page.extract_text()  # String
            list1 = line.split("\n")  # Lista com as Linhas
            for j, lines in enumerate(list1):
                if not j == len(list1) - 1:
                    list2 = list1[j].split()
                    if j > 0:
                        list3 = list1[j - 1].split()
                    if tipo == "eq":
                        if (list2[0][0] == "E" or list2[0][0] == "A") and list2[0][1:5].isnumeric():
                            Dados_Equip.append(lines_to_dic(list2, "eq"))
                        elif j > 0 and len(Dados_Equip) > 0 and list3[0][0] == "P" and list3[0][1:5].isnumeric():
                            Dados_Equip[-1]["Descrição"] = Dados_Equip[-1]["Descrição"] + " " + " ".join(list2)
                    elif tipo == "mo":
                        if list2[0][0] == "P" and list2[0][1:5].isnumeric():
                            Dados_MO.append(lines_to_dic(list2, "mo"))
                        elif j > 0 and len(Dados_MO) > 0 and list3[0][0] == "P" and list3[0][1:5].isnumeric():
                            Dados_MO[-1]["Descrição"] = Dados_MO[-1]["Descrição"] + " " + " ".join(list2)
                    elif tipo == "ma":
                        if list2[0][0] == "M" and list2[0][1].isnumeric():
                            Dados_Ma.append(lines_to_dic(list2, "ma"))
                        elif j > 0 and len(Dados_Ma) > 0 and list3[0][0] == "M" and list3[0][1:5].isnumeric():
                            Dados_Ma[-1]["Descrição"] = Dados_Ma[-1]["Descrição"] + " " + " ".join(list2)
                    elif tipo == "ccu":
                        if (len(list2[0]) == 7 or len(list2[0]) == 6) and list2[0][0:6].isnumeric():
                            Dados_CCU.append(lines_to_dic(list2, "ccu"))
                        elif j > 0 and len(Dados_CCU) > 0 and len(list3[0]) == 7 and list3[0][0:6].isnumeric():
                            Dados_CCU[-1]["Descrição"] = Dados_CCU[-1]["Descrição"] + " " + " ".join(list2)


# Função insere em cada planilha do excel os títulos das tabelas, baseando-se nas chaves utilizadas nos dicionários.

def titles(dic, planilha):
    for k, chave in enumerate(dic[0].keys()):
        titulo = planilha.cell(row=2, column=k + 2)
        titulo.value = chave


# Função insere os valores armazenados em suas respectivas planilhas no arquivo de Excel.

def cells(dic, planilha):
    for i, lines in enumerate(dic):
        for j, chave in enumerate(dic[0].keys()):
            celula = planilha.cell(row=i + 3, column=j + 2)
            celula.value = dic[i][chave]


# Recebe o diretório de cada PDF e onde será salvo o Excel final.

Dir_PDFs = input("Qual o diretório da pasta com os PDFs? ")
Dir_Eq = input("Qual o nome do arquivo de PDF de Equipamentos? ") + ".pdf"
Dir_MO = input("Qual o nome do arquivo de PDF de Mão de Obra? ") + ".pdf"
Dir_Ma = input("Qual o nome do arquivo de PDF de Materiais? ") + ".pdf"
Dir_CCU = input("Qual o nome do arquivo de PDF de CCUs? ") + ".pdf"
Dir_Excel = input("Qual o diretório em que o arquivo de Excel deve ser salvo? ")

# Listas que armazenarão os dicionários contendo as informações buscadas.

Dados_Equip = []
Dados_MO = []
Dados_Ma = []
Dados_CCU = []

# Aplicação da função "pdf_to_lines" para os quatro PDFs à serem lidos.

pdf_to_lines(Dir_Eq, "eq")
pdf_to_lines(Dir_MO, "mo")
pdf_to_lines(Dir_Ma, "ma")
pdf_to_lines(Dir_CCU, "ccu")

# Após os dados serem coletados, filtrados e armazenados em suas respectivas listas, eles serão agora distribuídos
# ordenadamente em um arquivo de Excel.

# Criação do arquivo de Excel e as planilhas que o comporão.

Sicro = Workbook()

Sicro_CCU = Sicro.active
Sicro_CCU.title = "CCU"
Sicro_Eq = Sicro.create_sheet("Equipamentos")
Sicro_MO = Sicro.create_sheet("Mão de Obra")
Sicro_Ma = Sicro.create_sheet("Materiais")

# Aplicação da função "titles" para as quatro planilhas à serem preenchidas.

titles(Dados_Equip, Sicro_Eq)
titles(Dados_MO, Sicro_MO)
titles(Dados_Ma, Sicro_Ma)
titles(Dados_CCU, Sicro_CCU)

# Aplicação da função "cells" para as quatro planilhas à serem preenchidas.

cells(Dados_Equip, Sicro_Eq)
cells(Dados_MO, Sicro_MO)
cells(Dados_Ma, Sicro_Ma)
cells(Dados_CCU, Sicro_CCU)

# Aqui o arquivo de excel é salvo no diretório desejado.

Sicro.save(os.path.join(Dir_Excel, 'Sicro.xlsx'))

# Visualização de final do código.

print("Pronto!")
